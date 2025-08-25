# em_sequence.py
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

#global definition
MAX_PC       = 64
PULSE_MAX    = 128
CWIDTH_SEQ   = 10
CWIDTH_DUMMY = 1
CWIDTH_PULSE = 3
fn_prefix = "em_pulse_seq"
en_debugging = 0

def search_closest(input_values):
    result = []

    for i, val in enumerate(input_values):
        min_lower_diff = float('inf')
        min_upper_diff = float('inf')
        lower_index = None
        upper_index = None

        for j, other in enumerate(input_values):
            if i == j:
                continue
            diff = abs(val - other)

            if other < val:
                if diff < min_lower_diff:
                    min_lower_diff = diff
                    lower_index = j
            elif other > val:
                if diff < min_upper_diff:
                    min_upper_diff = diff
                    upper_index = j

        result.append((lower_index, upper_index))
    return result

def calc_spreading(ps):
    spr, stdev = 10000.0, 1000.0
    #low_val, high_val = None, None
    idx_max = 0
    mean = float('inf')
    max_stdev = 0.0
    avg_stdev = 0.0
    sum_stdev = 0.0
    
    idx_sorted = [i for val, i in sorted((v, i) for i, v in enumerate(ps))]
    
    print(idx_sorted)

    length = len(ps)
    pnext = length + 1
    sum_squared = 0.0
    
    for i in range(2, length):
        temp = idx_sorted[0:i]
        temp.append(length)
        mean = (length - i) / i
        ret = search_closest(temp)
        sum_squared = 0.0
        temp_len = len(temp)
        avg_cnt = 0
        for idx, (low, high) in enumerate(ret):
            #if idx != 0 and idx != len(temp) - 1:
            if idx != 0:
                if temp[idx] > idx_max:
                    idx_max = temp[idx]
                avg_cnt += 1
                low_val = temp[low] if low is not None else None
                high_val = temp[high] if high is not None else None
                sum_squared += math.pow(mean - math.fabs(temp[idx]-temp[low] - 1), 2)
                if en_debugging == 1:
                    print(f"input[{idx}] = {temp[idx]} → "
                            f"smaller index: {low} (value: {low_val}), "
                            f"larger index: {high} (value: {high_val})")
        #sum_squared += math.pow(mean - math.fabs(temp[len(temp) - 1] - idx_max - 1), 2)
        variance = sum_squared / ((temp_len - 1) - 1)
        stdev = math.sqrt(variance)
        if stdev > max_stdev:
            max_stdev = stdev
        sum_stdev += stdev

        if en_debugging == 1:
            print("pc[%d] : m=%f, sum_squared=%f, v=%f, stdev=%f" \
                      %(i, mean, sum_squared, variance, stdev))
            print("************************************************************************************")
    avg_stdev = sum_stdev / avg_cnt
    return max_stdev, avg_stdev

def make_pseq_excel(ps):
    wb = Workbook()
    ws = wb.active
    
    idx_sorted = [i for val, i in sorted((v, i) for i, v in enumerate(ps))]

    fill_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    align = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style='thin', color='000000')
    border = Border(top=thin, left=thin, right=thin)
    
    for col in range(1, PULSE_MAX*2+1):
        if (col // 2) == 0:
            w = CWIDTH_SEQ
        elif (col % 2) == 1:
            w = CWIDTH_DUMMY
        else:
            w = CWIDTH_PULSE
        ws.column_dimensions[get_column_letter(col)].width = w
    
    ws['A1'] = "sequence"
    
    for i in range(len(ps)):
        col = 2 * i + 2
        ws.cell(row=1, column=col, value=i + 1)
        
    last_value_in_col = {}
    
    for seq_num, col_idx in enumerate(idx_sorted):
        row = 2 * seq_num + 2
        col = 2 * col_idx + 2
        ws.cell(row=row, column=1, value=seq_num + 1)
    
        if col not in last_value_in_col:
            cell = ws.cell(row=row, column=col)
            cell.value = seq_num+1
            cell.fill = fill_color
            cell.alignment = align
            cell.border = border
            last_value_in_col[col] = cell.value
            for sn in range(seq_num):
                r = 2 * seq_num + 2
                c = 2 * idx_sorted[sn] + 2
            
                cell = ws.cell(row=r, column=c)
                cell.value = sn+1
                cell.alignment = align
                cell.border = border
        else:
            print("%%%%%%%%%%%%%%%%%% case error case %%%%%%%%%%%%%%%%%%")
            return -1
    
    max_row = 2 * len(idx_sorted) + 1
    for row in range(2, max_row + 1, 2):
        for col in range(2, len(ps)*2+2, 2):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                cell.border = Border(bottom=thin)

            cell_right = ws.cell(row=row, column=col + 1)
            if cell_right.value is None:
                cell_right.border = Border(bottom=thin)
    
    wb.save(f"em_pulse_seq_{len(ps)}.xlsx")
    return 0
# end def make_pseq_excel()
    
def non_recursive_seq(pulse_count):
    sequence = [1]
    next_sequence = []
    while len(sequence) < pulse_count:
        next_sequence.clear()
        for num in sequence:
            odd = 2 * num - 1
            if odd <= pulse_count:
                next_sequence.append(odd)
        for num in sequence:
            even = 2 * num
            if even <= pulse_count:
                next_sequence.append(even)
        sequence = next_sequence[:]

    return sequence

def main():
    try:
        ef = Workbook()
        sheet = ef.active
        
        while True:
            inStr = input("\nEnter the number of pulses (e.g. 32): ")
            if inStr == 'exit':
                break
            pc = int(inStr)
            if pc <= 0 or pc > MAX_PC:
                print("\nWrong pulse count! 0 < number < %d" % (MAX_PC+1))
            else:
                result = non_recursive_seq(pc)
                print("Generated sequence:")
                print(result)
                if len(result) > 2:
                    maxs, avgs = calc_spreading(result)
                    print("Flicker rate : max(%f), avg(%f)" %(maxs, avgs))
                ret = make_pseq_excel(result)
                print("Error(%d)" %(ret)) if ret == -1 else print(f"A pulse sequence diagram was generated and saved as ./{fn_prefix}_{len(result)}.xlsx")
                
    except ValueError:
        print("Please enter a valid integer.")

if __name__ == "__main__":
    main()
